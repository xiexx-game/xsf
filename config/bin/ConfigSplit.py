# -*- coding: utf-8 -*-  

import codecs
# 防止中文乱码
import importlib
import os
import sys

import openpyxl

importlib.reload(sys)

# 分割符
C_SPACE = ","
# 结束符
C_END = "\n"


IN_PATH = u"../../"
OUT_PATH = sys.argv[1]
OUT_SERVER_PATH = sys.argv[2]
OUT_CPP_PATH = sys.argv[3]

# 获取指定后缀文件
def getFile(dir, suffix):
    res = []
    for root, directory, files in os.walk(dir):
        for filename in files:
            name, suf = os.path.splitext(filename)
            if suf == suffix:
                res.append(os.path.join(root, filename))
    return res


# 打开excel
def open_excel(file):
    try:
        wb = openpyxl.load_workbook(file)
        #sheet_name = wb.sheetnames[0]
        #return wb[sheet_name]
        return wb
    except Exception as e:
        print("open_excel catch exception:" + str(e))


# 根据索引获取Excel表格中的数据 参数:file：Excel文件路径, colnameindex：表头列名所在行的索引, by_index：表的索引
def excel_table_byindex(sheet_data, order):
    nrows = sheet_data.max_row  # 行数
    ncols = sheet_data.max_column  # 列数
    rowlist = []

    #print("excel_table_byindex", nrows, ncols)

    for rowIndex in range(3, nrows+1):
        if sheet_data.cell(rowIndex, 1).value != "*":     # 该行未标记，不导出
            continue
        
        collist = []
        for colIndex in range(2, ncols+1):
            cell_value = sheet_data.cell(rowIndex, colIndex).value

            #print("flag=", colIndex, order, sheet_data.cell(order, colIndex).value)
            if sheet_data.cell(order, colIndex).value == "*":     # 该列标记，才导出
                if rowIndex >= 6:
                    #print(rowIndex, colIndex, cell_value)
                    if cell_value is None:
                        collist.append("")
                    else:
                        csv_data_type = sheet_data.cell(4, colIndex).value
                        if csv_data_type == "uint" or csv_data_type == "int" or csv_data_type == "ulong":       # 如果是整形
                            collist.append(str(cell_value))
                        elif csv_data_type == "float":
                            x_new = '{:.3f}'.format(float(cell_value))
                            collist.append(str(x_new))
                        else:
                            collist.append(str(cell_value))
                else:
                    collist.append(str(cell_value))

        if len(collist) > 0:
                rowlist.append(collist)

    return rowlist


# 保存csv文件
def savaToCSV(_file, _list, _path):
    print(u"savaToCSV:" + _file + " " + _path)
    filename = ""
    content = ""
    # 生成文件内容
    for collist in _list:
        line = ""
        for col in collist:
            if len(line) <= 0:
                line = col
            else:
                line += C_SPACE + col 

        content += line + C_END

    # 生成文件后缀
    fname = os.path.basename(_file).split('.')
    filename = fname[0] + ".csv"

    # 写文件
    if len(filename) > 0 and len(content) > 0:
        filename = pathCombine(_path, filename)
        print(u"输出文件:" + filename)
        file_object = codecs.open(filename, 'w', "utf-8")
        file_object.write(content)
        file_object.close()


def main():
    filelist = getFile(IN_PATH, ".xlsx")
    if len(filelist) > 0:
        path = ""
        # 遍历文件生成csv
        for file in filelist:
            excel_data = open_excel(file)
            sheet_name = excel_data.sheetnames[0]
            #return wb[sheet_name]
            print("表单：", sheet_name)
            sheet_data = excel_data[sheet_name]
            datalist_Server = excel_table_byindex(sheet_data, 1)
            datalist_Client = excel_table_byindex(sheet_data, 2)
            if len(datalist_Client) > 0 and os.path.exists(OUT_PATH):
                savaToCSV(sheet_name, datalist_Client, OUT_PATH)
            if len(datalist_Server) > 0 and os.path.exists(OUT_SERVER_PATH):
                savaToCSV(sheet_name, datalist_Server, OUT_SERVER_PATH)
            if len(datalist_Server) > 0 and os.path.exists(OUT_CPP_PATH):
                savaToCSV(sheet_name, datalist_Server, OUT_CPP_PATH)
    else:
        print(u"没有找到任何excel文件！")


def pathCombine(path1: str, path2: str):
    if path1.endswith("/") or path1.endswith("\\"):
        return path1 + path2
    return path1 + "/" + path2


if __name__ == "__main__":
    main()
